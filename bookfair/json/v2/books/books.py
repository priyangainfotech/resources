#!/usr/bin/env python3
import sys

from openpyxl import load_workbook
excel_book= sys.argv[1]
json_file = sys.argv[2]
wb = load_workbook(excel_book)
from json import dumps
import sys

# get sheet by name
sheet = wb["Sheet1"]

# get number of rows
rows = sheet.max_row

# get number of columns
columns = sheet.max_column


# list to store all the rows of excel file as dictionary
books = []
for i in range(1, rows):
    row = {}
    for j in range(1, columns):
        column_name = sheet.cell(row=1, column=j)
        row_data = sheet.cell(row=i+1, column=j)
        if(column_name.value == "code"):
            row.update(
            {
                "code":  row_data.value
            })
        elif column_name.value == "title":
            row.update(
            {
                "title": row_data.value
            })
        elif column_name.value =="isbn":
            row.update(
            {
                "isbn": row_data.value
            }) 
        elif column_name.value == "price":
            row.update(
            {
                "price": row_data.value
            }) 
        elif column_name.value == "classification":
            row.update(
            {
                "classification": row_data.value
            }) 
            
        elif column_name.value == "author":
            row.update(
            {
                "author": row_data.value
            })
       
    books.append(row)

# convert into json
json_data = dumps(books)
print(json_data)


out_file  = open (json_file, "w")
out_file.write(json_data)
out_file.close()