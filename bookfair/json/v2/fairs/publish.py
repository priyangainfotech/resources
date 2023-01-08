#!/usr/bin/env python3
import sys

from openpyxl import load_workbook
excel_book= sys.argv[1]
json_file = sys.argv[2]
wb = load_workbook(excel_book)
from json import dumps
import sys

def capitalize(string):
    if isinstance(string, str):
        words = string.split(" ") # just change the split(" ") method
        return ' '.join([word.capitalize() for word in words])
    return string


# get sheet by name
sheet = wb["Sheet1"]

# get number of rows
rows = sheet.max_row

# get number of columns
columns = sheet.max_column

# list to store all the rows of excel file as dictionary
books = []
for i in range(1, rows):
    row = {}
    for j in range(1, columns):
        column_name = sheet.cell(row=1, column=j)
        row_data = sheet.cell(row=i+1, column=j)
        print(column_name.value)
        if(column_name.value == "stall_no"):
            sno = str(row_data.value)
            row.update(
            {
                "stall_no":  [x.strip() for x in sno.split(',')]
            })
        else:
            row.update({
            column_name.value: capitalize(row_data.value)
              })
       
    books.append(row)

# convert into json
json_data = dumps(books)
#print(json_data)


out_file  = open (json_file, "w")
out_file.write(json_data)
out_file.close()