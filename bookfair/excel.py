#!/usr/bin/env python3

from openpyxl import load_workbook
wb = load_workbook("book_list.xlsx")
from json import dumps
import sys

# get sheet by name
sheet = wb["Sheet1"]

# get number of rows
rows = sheet.max_row

# get number of columns
columns = sheet.max_column

alp = "1,2,3"


# list to store all the rows of excel file as dictionary
root = {}
stalls = []
for i in range(1, rows):
    publication = {}
    row = {}
    for j in range(1, columns):
        column_name = sheet.cell(row=1, column=j)
        row_data = sheet.cell(row=i+1, column=j)
        if(column_name.value == "STALL NO"):
            sno = str(row_data.value)
            row.update(
            {
                "stall_no":  [x.strip() for x in sno.split(',')]
            })
    
    
        elif column_name.value =="NAME":
            row.update(
            {
                "name": row_data.value
            })
        elif column_name.value =="S.No":
            row.update(
            {
                "id": row_data.value
            }) 
        else:
            row.update(
            {
                column_name.value: row_data.value
            })

        publication.update({"publications":row})
    stalls.append(publication)

root.update({"stalls":stalls})

# convert into json
json_data = dumps(root)
print(json_data)


out_file  = open ("output.json", "w")
out_file.write(json_data)
out_file.close()